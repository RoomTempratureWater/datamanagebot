import openpyxl


def append(username,rank):
    row = [username,rank]
    file = 'Data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['accounts']
    ws.append(row)
    wb.save(file)

    ws = wb['accounts_meta']
    if rank[0] == 'i':
        new = int(ws['A1'].value) + 1
        ws['A1'] = new
    elif rank[0] == 'b':
        new = int(ws['B1'].value) + 1
        ws['B1'] = new
    elif rank[0] == 's':
        new = int(ws['C1'].value) + 1
        ws['C1'] = new
    elif rank[0] == 'g':
        new = int(ws['D1'].value) + 1
        ws['D1'] = new
    elif rank[0] == 'p':
        new = int(ws['E1'].value) + 1
        ws['E1'] = new
    elif rank[0] == 'd':
        new = int(ws['F1'].value) + 1
        ws['F1'] = new
    else:
        print(rank[0])
        new = int(ws['G1'].value) + 1
        ws['G1'] = new
    wb.save(file)

#test for saving data into file func
'''username = 'suckmyaz'
rank = d2
append(username,rank)'''

def getdata():
    i,b,s,g,p,d,im = 0,0,0,0,0,0,0
    file = 'Data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['accounts_meta']
    i = ws['A1'].value
    b = ws['B1'].value
    s = ws['C1'].value
    g = ws['D1'].value
    p = ws['E1'].value
    d = ws['F1'].value
    im = ws['G1'].value
    str = "iron = {}\nbronze = {}\nsilver = {}\ngold = {}\nplat = {}\ndiamond = {}\nimmortal = {}".format(i,b,s,g,p,d,im)
    return str

#print(getdata()) ---test to check getdata func

def del_acc(username, rank):
    file = 'Data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['accounts']
    i = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            s = ws.cell(r, c).value
            if s != None and username in s:
                ws.cell(r, c).value = None
                ws.cell(r, c+1).value = None
                #print("row {} col {} : {}".format(r, c, s))
                i += 1
    wb.save(file)

    ws = wb['accounts_meta']
    if rank[0] == 'i':
        new = int(ws['A1'].value) - 1
        ws['A1'] = new
    elif rank[0] == 'b':
        new = int(ws['B1'].value) - 1
        ws['B1'] = new
    elif rank[0] == 's':
        new = int(ws['C1'].value) - 1
        ws['C1'] = new
    elif rank[0] == 'g':
        new = int(ws['D1'].value) - 1
        ws['D1'] = new
    elif rank[0] == 'p':
        new = int(ws['E1'].value) - 1
        ws['E1'] = new
    elif rank[0] == 'd':
        new = int(ws['F1'].value) - 1
        ws['F1'] = new
    else:
        print(rank[0])
        new = int(ws['G1'].value) - 1
        ws['G1'] = new
    wb.save(file)

#del_acc('suckmyaz', 'd2')

def get_all():
    file = 'Data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['accounts']
    accounts_data_str = ''
    i = 0
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        rank = ws.cell(r, 2).value
        accounts_data_str += '{} - {}\n'.format(name,rank)

    return accounts_data_str

#print(get_all())